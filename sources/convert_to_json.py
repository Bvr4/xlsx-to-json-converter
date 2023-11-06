import openpyxl
import json
import datetime

def convert_to_json(fichier_entree, chemin_sortie):
    document = openpyxl.load_workbook(fichier_entree, data_only=True)

    feuille_travaux = document["Travaux"]
    feuille_financeurs = document["Financeurs"]

    financeurs = lister_financeurs(feuille_financeurs)
    travaux = lister_travaux(feuille_travaux, financeurs)

    return ecrire_fichier_json(chemin_sortie, travaux)


def lister_financeurs(feuille):
    # Création de la liste des financeurs (vide)
    liste_financeurs = []
    # On parcourt les données financeurs :
    i = 2
    while feuille.cell(i, 1).value != None:
        financeur = {}
        financeur['Numéro Travaux'] = feuille.cell(i, 1).value
        financeur['Financeur'] = feuille.cell(i, 2).value
        financeur['Type financeur'] = feuille.cell(i, 3).value[:30]
        financeur['Montant HT participation'] = f"{(feuille.cell(i, 4).value):.2f}"
        financeur['Taux participation'] = str(feuille.cell(i, 5).value)
        liste_financeurs.append(financeur)
        i += 1
    return liste_financeurs


def lister_travaux(feuille, financeurs):
    # Création de la liste des travaux (vide)
    travaux = []
    # On parcourt les données travaux :
    i = 2
    while feuille.cell(i, 1).value != None:
        ligne_travaux = {}
        ligne_travaux['codeTravaux'] = feuille.cell(i, 1).value
        ligne_travaux['insee'] = feuille.cell(i, 2).value
        ligne_travaux['nomCommune'] = feuille.cell(i, 3).value
        if feuille.cell(i, 4).value !=  0:
            ligne_travaux['dateCommande'] = feuille.cell(i, 4).value.strftime("%Y-%m-%d")
        else:
            ligne_travaux['dateCommande'] = ""
        if feuille.cell(i, 5).value !=  0:
            ligne_travaux['dateReception'] = feuille.cell(i, 5).value.strftime("%Y-%m-%d")
        else:
            ligne_travaux['dateReception'] = ""
        ligne_travaux['typeTravaux'] = feuille.cell(i, 6).value
        ligne_travaux['montantTravaux']=f"{(feuille.cell(i, 7).value):.2f}"

        # recherche des financeurs
        autresFinanceurs=[]
        for financeur in financeurs:
            if financeur['Numéro Travaux'] == ligne_travaux['codeTravaux']:
                autreFinanceur = {}
                autreFinanceur['typeFinanceur']=financeur['Type financeur']
                autreFinanceur['identite']=financeur['Financeur']
                autreFinanceur['montantParticipation']=financeur['Montant HT participation']
                autreFinanceur['tauxParticipation']=financeur['Taux participation']
                autresFinanceurs.append(autreFinanceur)
        
        # Si on a trouvé des financeurs pour ce projet, on crée la variable
        if len(autresFinanceurs)>0:
            ligne_travaux['autresFinanceurs']=autresFinanceurs

        travaux.append(ligne_travaux)
        i += 1 
    
    return travaux


def ecrire_fichier_json(chemin_sortie, donnees):
    # génération d'un nom de fichier unique
    now = datetime.datetime.now()
    nom_fichier = "export_" + str(now.year) + "-" + str(now.month) + "-" + str(now.day) + "_" + str(now.hour) + "-" + str(now.minute) + "-" + str(now.second) + ".json"
    chemin_fichier = chemin_sortie + "/" + nom_fichier

    # écriture dans un fichier JSON
    with open(chemin_fichier, "w") as fichier:
        json.dump(donnees, fichier)
    
    return nom_fichier
    